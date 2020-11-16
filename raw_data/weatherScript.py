import openpyxl
import datetime

datasetWB = openpyxl.load_workbook(filename="dataset.xlsx", read_only=True)
datasetSheet = datasetWB.active


data = []#[[]] * 207623


weatherData = [ 'humidity', 'pressure',	'temperature', 'weather_description', 'wind_direction', 'wind_speed']

holidays2015 = [datetime.datetime(2015, 1, 1), datetime.datetime(2015, 1, 19), datetime.datetime(2015, 2, 4), datetime.datetime(2015, 2, 16), 
                datetime.datetime(2015, 3, 31), datetime.datetime(2015, 5, 25), datetime.datetime(2015, 6, 19), datetime.datetime(2015, 7, 3),
                datetime.datetime(2015, 7, 4), datetime.datetime(2015, 9, 7), datetime.datetime(2015, 11, 11), datetime.datetime(2015, 11, 26),
                datetime.datetime(2015, 12, 25)]

holidays2016 = [datetime.datetime(2016, 1, 1), datetime.datetime(2016, 1, 18), datetime.datetime(2016, 2, 4), datetime.datetime(2016, 2, 15), 
                datetime.datetime(2016, 3, 31), datetime.datetime(2016, 5, 30), datetime.datetime(2016, 7, 19),
                datetime.datetime(2016, 7, 4), datetime.datetime(2016, 9, 5), datetime.datetime(2016, 11, 11), datetime.datetime(2016, 12, 25),
                datetime.datetime(2016, 12, 26)]
#weatherData = [ 'humidity']

def copyWorkbook(oldSheet):

    for row in oldSheet.rows:
        temp = []
        for cell in row:
            temp.append(cell.value)
        data.append(temp)
       # for cell in row:
         #   print(cell.value)
            
            #openpyxl.cell.WriteOnlyCell()
            #newSheet.cell(row=counterRow, column=counterCol).value = cell.value
            #counterCol += 1
    #counterCol = 1
    #counter += 1


def getFirstEmptyColumn(sheet):
    index = 1
    while sheet.cell(row=1, column=index).value:
        index += 1
    return index


def getLAColumn(sheet):
    index = 1
    while sheet.cell(row=1, column=index).value != 'Los Angeles':
        index += 1
    return index

def getDateColumn(sheet):
    index = 1
    while sheet.cell(row=1, column=index).value != 'datetime' and sheet.cell(row=1, column=index).value != 'Date':
        index += 1
    return index

def getIndexOfDate(sheet, date, dateIndex):
    return 19718
    #return 37261
    index = 37237
    while sheet.cell(row=index, column=dateIndex).value != date:
      #  if index % 100 == 0: 
        print(index)
        print( sheet.cell(row=index, column=dateIndex).value)
        index += 1
    return index

def getLastRow(sheet):
    index = 1
    while sheet.cell(row=index, column=1).value:
        index += 1
    return index


def extractDayAndMonth():
    data[0].append("Day")
    data[0].append("Month")
    for i in range(1, 207649):
        date = data[i][0]
        data[i].append(date.weekday())
        data[i].append(date.strftime("%m") )

def validateDate():
    allDates = []
    for i in range(3, len(data)):
        if data[i][0] != data[i - 1][0]:
            allDates.append(data[i-1][0])

    excelFileName = 'incident-dataset.xlsx'
    workbook = openpyxl.load_workbook(filename=excelFileName)
    incidentSheet = workbook.active
    indexOfRowDataSet = 0

    for row in (incidentSheet.rows):
        for cell in row[:1]:
            date_time_obj = datetime.datetime.strptime(cell.value, '%m/%d/%Y')
            print(indexOfRowDataSet)
            if date_time_obj.year != allDates[indexOfRowDataSet].year and date_time_obj.month != allDates[indexOfRowDataSet].month and date_time_obj.day != allDates[indexOfRowDataSet].day:
                print(cell.value)
                print(allDates[indexOfRowDataSet])
                raise "ERROR"
            indexOfRowDataSet += 1

    print(allDates)

def addHolidays():
    data[0].append("Holiday")
    for i in range(1, 207649):
        date = (data[i][0]).replace(hour=0, minute=0)
        if date in holidays2015 or date in holidays2016:
            data[i].append(1)
        else:
            data[i].append(0)
        
  
    
    

def addIncidents():
    data[0].append("Incidents")
    excelFileName = 'incident-dataset.xlsx'
    workbook = openpyxl.load_workbook(filename=excelFileName)
    incidentSheet = workbook.active
    indexOfRowDataSet = 1

    #curDateTime = date_time_obj
    hour = 0
    for row in (incidentSheet.rows):
        if row[0].value == 'Time':
            continue

        dateOfIncidentSheet = row[0]

        date_time_obj = datetime.datetime.strptime(dateOfIncidentSheet.value, '%m/%d/%Y')
        date_time_obj += datetime.timedelta(hours=0)
        
        for cellToCopy in row[1:]:
            for i in range(12):
                date = data[indexOfRowDataSet][0]
                if  date_time_obj == date:
                  #  print('    Added')
                    if cellToCopy.value:
                        data[indexOfRowDataSet].append(cellToCopy.value)
                        
                    else:
                        data[indexOfRowDataSet].append(0)
                    indexOfRowDataSet += 1
              #  else:
                  #  print('    DateTime is missing: ', date_time_obj)
                   # return
                date_time_obj += datetime.timedelta(minutes=5)
          #  date_time_obj += datetime.timedelta(hours=1)
                
def buildMaxLaneClosure():
    excelFileName = 'lane_closures_dataset.xlsx'
    workbook = openpyxl.load_workbook(filename=excelFileName)
    workSheet = workbook.active
    startDateCol = 15
    endDateCol = 18
    closureLanesCol = 48
    totalLanesCol = 49

    maxRow = 8805
    currStartTimeRow = 2
    currMaxLanes = 0

    currEndTimeRow = 2

   # currStartTime = datetime.datetime.strptime(workSheet.cell(row=currStartTimeRow, column=startDateCol).value, '%Y-%m-%d %H:%M:%S')
   # currStartTime = workSheet.cell(row=currStartTimeRow, column=startDateCol).value
   # currMaxLanes = workSheet.cell(row=currStartTimeRow, column=closureLanesCol).value / workSheet.cell(row=currStartTimeRow, column=totalLanesCol).value
    
    currStartTimeRow += 1

    #currEndTime = workSheet.cell(row=currEndTimeRow, column=endDateCol).value

    #currEndTime = datetime.datetime.strptime(workSheet.cell(row=currEndTimeRow, column=endDateCol).value, '%Y-%m-%d %H:%M:%S')
    
    '''
    maxLaneClosuresStartTime = []

    while currStartTimeRow <= maxRow:
        tempStartTime = workSheet.cell(row=currStartTimeRow, column=startDateCol).value
       # tempStartTime= datetime.datetime.strptime(workSheet.cell(row=currStartTimeRow, column=startDateCol).value, '%Y-%m-%d %H:%M:%S')
        tempMax = workSheet.cell(row=currStartTimeRow, column=closureLanesCol).value / workSheet.cell(row=currStartTimeRow, column=totalLanesCol).value
      #  print('Comparing ', tempStartTime, ' with ', currStartTime)
      #  print('Result: ', tempStartTime > currStartTime )
      #  print('Curr max lanes: ', currMaxLanes)
        if tempStartTime > currStartTime:
            maxLaneClosuresStartTime.append((currStartTime, currMaxLanes))
            currMaxLanes = tempMax
            currStartTime = tempStartTime
        else: 
            currMaxLanes =  max(currMaxLanes, tempMax)
          #  print(currStartTime, ' ', maxRow)
            if currStartTimeRow == maxRow:
                maxLaneClosuresStartTime.append((tempStartTime, currMaxLanes))
        currStartTimeRow += 1
    '''

  #  print('#########################################################################')
    #print ('Here')
  #  for i in range(len(maxLaneClosuresStartTime)):
  #      print(maxLaneClosuresStartTime[i])

   # print('#########################################################################')

    currMaxLanes = 0

   # currEndTimeRow = 2

    #currEndTime = workSheet.cell(row=currEndTimeRow, column=endDateCol).value
   
   # currEndTimeRow += 1
    startTimeList = []   


    currStartTimeRow = 2
    while currStartTimeRow <= maxRow:
        #startTime = datetime.datetime.strptime(workSheet.cell(row=currStartTimeRow, column=startDateCol).value, '%Y-%m-%d %H:%M:%S')
        startTime = workSheet.cell(row=currStartTimeRow, column=startDateCol).value
        lanesClosed = workSheet.cell(row=currStartTimeRow, column=closureLanesCol).value
        totalLanes = workSheet.cell(row=currStartTimeRow, column=totalLanesCol).value
        if lanesClosed != 0 and totalLanes >= 3:
            lanesClosedPercent = lanesClosed / totalLanes
            startTimeList.append((startTime, lanesClosedPercent))

        currStartTimeRow += 1

    startTimeList.sort(key=lambda tup: tup[0])  # sorts in place

   # currEndTime = workSheet.cell(row=currEndTimeRow, column=endDateCol).value
    endTimeList = []   


    currEndTimeRow = 2
    while currEndTimeRow <= maxRow:
        endTime = datetime.datetime.strptime(workSheet.cell(row=currEndTimeRow, column=endDateCol).value, '%Y-%m-%d %H:%M:%S')
        lanesClosed = workSheet.cell(row=currEndTimeRow, column=closureLanesCol).value 
        totalLanes = workSheet.cell(row=currEndTimeRow, column=totalLanesCol).value
        if lanesClosed != 0 and totalLanes >= 3:
            lanesClosedPercent = lanesClosed / totalLanes
            endTimeList.append((endTime, -1 * lanesClosedPercent))
        currEndTimeRow += 1

    endTimeList.sort(key=lambda tup: tup[0])  # sorts in place

  #  print('Edge case check: ', endTimeList[len(endTimeList) - 1])



   # maxLaneClosuresEndTime.sort(key=lambda tup: tup[0])  # sorts in place
    


    lengthEndTime = len(endTimeList)
    lengthStartTime = len(startTimeList)

    currStartTime = startTimeList[0][0]
    currEndTime = endTimeList[0][0]

    startTimeCounter = 1
    endTimeCounter = 1
    
    result = []

    for temp in startTimeList:
        result.append(temp)
    for temp in endTimeList:
        result.append(temp)
        
    result.sort(key=lambda tup: tup[0])

    #print('LENGTH OF ENDTIME: ', len(endTimeList))
    #print('LENGTH OF STARTTIME: ', len(startTimeList))
   # print('LENGTH OF RESULT: ', len(result))
   # for i in range(len(result)):
  #      print(result[i][0].strftime("%m/%d/%Y, %H:%M:%S"), ' ' , result[i][1])
  #  print('#########################################################################')

    data[0].append('LaneClosures')

   # lanesClosed = len(result)
    lanesClosedCounter = 0
  #  currLanesClosed = 0
    lastRowIndex = 207649

    stack = []
    stack.append(0)
    for i in range(1, lastRowIndex):
        date = data[i][0]
      #  print('Date checking: ', date)
       # print(result[lanesClosedCounter + 1][0].strftime("%m/%d/%Y, %H:%M:%S"), ' ' , result[lanesClosedCounter + 1][1])
       # print('stack is: ', stack)
        

#index < len

        if lanesClosedCounter + 1 >= len(result):
            stack = [0]

        elif date >= result[lanesClosedCounter][0]:
         #   print(result[lanesClosedCounter][0].strftime("%m/%d/%Y, %H:%M:%S"), ' ' , result[lanesClosedCounter][1])
            lanesClosedCounter = updateSet(lanesClosedCounter, stack, result, date)



        data[i].append(max(stack))    


    #    data[i].append(currLanesClosed)



   # for i in range(len(maxLaneClosures)):
       # print(maxLaneClosures[i])



def updateSet(lanesClosedCounter, stack, result, date):
    while lanesClosedCounter < len(result) and (date >= result[lanesClosedCounter][0]):
     #   if date >= result[lanesClosedCounter + 1][0]:
        #    lanesClosedCounter += 1

        tempLanesClosed = result[lanesClosedCounter][1]
        # print('Here: ' , tempLanesClosed)
        if tempLanesClosed < 0:
          #  if (-1*tempLanesClosed) in stack:
            stack.remove(-1*tempLanesClosed)
        elif tempLanesClosed > 0 :
            stack.append(tempLanesClosed)
        lanesClosedCounter += 1
    return lanesClosedCounter

def convertToEpoch():
    for i in range(1, len(data)):
        if (data[i][0] != None):
            data[i][0] = data[i][0].timestamp()


def addWeather():
    lastRowIndex = 207649
   # print(lastRowIndex)
    last_date_time = datasetSheet.cell(row=lastRowIndex,column=1).value

    for s in weatherData:
        indexOfRowDataSet = 0
        #   emptyColumnIndex = getFirstEmptyColumn(datasetSheet)
        excelFileName = 'weather/' + s + '.xlsx'
        workbook = openpyxl.load_workbook(filename=excelFileName)
        

        sheetToCopyFrom = workbook.active
        #sheetToCopyFrom.reset_dimensions()
        LAcolIndex = getLAColumn(sheetToCopyFrom)
        #weatherSheetDateIndex = getDateColumn(sheetToCopyFrom)
       # indexOfDate = getIndexOfDate(sheetToCopyFrom, date_time_obj, weatherSheetDateIndex)
        indexOfDate = 19718

        print('Copying data for: ', s)

        data[indexOfRowDataSet].append(s)
        #newSheet.cell(row=indexOfRowDataSet, column=emptyColumnIndex).value = s
        indexOfRowDataSet += 1

        #   currRowIndex = 
        date_time_obj = sheetToCopyFrom.cell(row=indexOfDate, column=1).value
        
     #   date_time_obj = datetime.datetime.strptime(startCell.value, '%m/%d/%Y')
# date_time_obj.replace(hour=0, minute=0)
        
        while date_time_obj <= last_date_time:
            
            weatherCell = sheetToCopyFrom.cell(row=indexOfDate, column=LAcolIndex)
            for i in range(12):
                date = data[indexOfRowDataSet][0]
                if  date_time_obj == date:
                  #  print('    Added')
                    if not weatherCell.value is None:
                        data[indexOfRowDataSet].append(weatherCell.value)
                        
                    else:
                        data[indexOfRowDataSet].append(-1)
                    indexOfRowDataSet += 1
               # else:
                 #   print('    DateTime is missing: ', date_time_obj)
                   # return
                date_time_obj += datetime.timedelta(minutes=5)
            indexOfDate += 1
          #  date_time_obj += datetime.timedelta(hours=1)

#def trimDataSet(startTime, endTime):
#    while

def addWeatherToDataSet(sheet):

    print('Getting date column')
    datasetDateColumn = getDateColumn(sheet)

    date_time_obj = datasetSheet.cell(row=2,column=datasetDateColumn).value
    print('Getting last row')

    lastRowIndex = 207623
   # print(lastRowIndex)

    print('Done getting last row')
    last_date_time = datasetSheet.cell(row=lastRowIndex,column=datasetDateColumn).value

    emptyColumnIndex = getFirstEmptyColumn(datasetSheet)

    #newWorkbook = openpyxl.Workbook(write_only = True)  

    #newSheet = newWorkbook.active

    copyWorkbook(datasetSheet)
    #validateDate()
    extractDayAndMonth()
    addHolidays()
    addIncidents()
    addWeather()
    buildMaxLaneClosure()
    convertToEpoch()
    datasetWB.close()
    '''
  #  date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')

    print('First Date-time for dataSet:', date_time_obj)
    print('Last Date-time for dataSet:', last_date_time)
    for s in weatherData:
        indexOfRowDataSet = 0
     #   emptyColumnIndex = getFirstEmptyColumn(datasetSheet)
        excelFileName = 'weather/' + s + '.xlsx'
        workbook = openpyxl.load_workbook(filename=excelFileName)
        

        sheetToCopyFrom = workbook.active
        #sheetToCopyFrom.reset_dimensions()
        print('Here')
        LAcolIndex = getLAColumn(sheetToCopyFrom)
        print('Here')
        weatherSheetDateIndex = getDateColumn(sheetToCopyFrom)
        print('Here')
        indexOfDate = getIndexOfDate(sheetToCopyFrom, date_time_obj, weatherSheetDateIndex)
        print('Copying data for: ', s)

        data[indexOfRowDataSet].append(s)
        #newSheet.cell(row=indexOfRowDataSet, column=emptyColumnIndex).value = s
        indexOfRowDataSet += 1

        curDateTime = date_time_obj
     #   currRowIndex = 
        while curDateTime <= last_date_time:
           # print(curDateTime)
            cellToCopy = sheetToCopyFrom.cell(row=indexOfDate, column=LAcolIndex)
            for i in range(12):
                data[indexOfRowDataSet].append(cellToCopy.value)
               # newSheet.cell(row=indexOfRowDataSet, column=emptyColumnIndex).value = sheetToCopyFrom.cell(row=indexOfDate, column=LAcolIndex).value
                indexOfRowDataSet += 1

           # print('Copied data for hour: ', curDateTime, ' from cell: (' ,indexOfDate, ',' , LAcolIndex ,')' )
            indexOfDate += 1
            curDateTime = sheetToCopyFrom.cell(row=indexOfDate, column=weatherSheetDateIndex).value
           # print(curDateTime)
       # emptyColumnIndex += 1
        workbook.close()
    print(data[0])
    print(data[1])
    print(data[2])
    '''
    newWorkbook = openpyxl.Workbook(write_only = True)  
    newSheet = newWorkbook.create_sheet()

    print('Saving workbook')
    for row in data:
        newSheet.append(row)

    newWorkbook.save('dataset1.xlsx')
   # datasetWB.defined_names.definedName = []
    





addWeatherToDataSet(datasetSheet)
#buildMaxLaneClosure()
